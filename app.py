"""
Johnsons Books — Publisher File Extractor
Streamlit app that:
 - Accepts upload of multiple publisher Excel files (all structurally different)
 - Auto-detects the right parser for each file based on filename
 - Cross-references prices against Prezzi.xlsx
 - Exports a unified output.xlsx with 8 columns
"""

import io
import json
import re
import time
import urllib.request

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PREZZI_PATH = "/Users/filippopalma/Desktop/AI/Johnsons Books/Prezzi/Prezzi.xlsx"

HEADERS_BASE = [
    "EAN", "TITOLO", "EDITORE", "PAESE", "SCONTO",
    "Prezzo Editore",
    "Prezzo al pubblico consigliato (IVA inclusa)",
    "Costo netto di cessione ad Amazon",
]

HEADERS_WEB = [
    "Descrizione prodotto",
    "Numero di pagine",
    "Rilegatura",
    "Lingua",
    "Data di pubblicazione",
    "Codice soggetto (BISAC)",
    "Dimensione L (mm)",
    "Dimensione W (mm)",
    "Dimensione H (mm)",
    "Peso (kg)",
]

HEADERS_FIXED = [
    "Codice fornitore",
    "Tipo di prodotto",
    "Ruolo collaboratore",
    "Tipo di lingua",
    "Rischio GDPR",
    "Batterie necessarie",
    "Merci pericolose",
    "Avviso et\u00e0 DSG UE",
    "Avviso DSG UE",
]

HEADERS = HEADERS_BASE  # alias — usato da process_files per output base

FIXED_VALUES = {
    "Codice fornitore":  "Johnsons Books",
    "Tipo di prodotto":  "ABIS_BOOK",
    "Ruolo collaboratore": "Autore",
    "Tipo di lingua":    "Pubblicato",
    "Rischio GDPR":      "Nessuna informazione elettronica memorizzata",
    "Batterie necessarie": "No",
    "Merci pericolose":  "Non applicabile",
    "Avviso et\u00e0 DSG UE": "Nessun avvertimento applicabile",
    "Avviso DSG UE":     "Nessun avvertimento applicabile",
}

# Auto-detection keywords
PRICE_KW     = ["price", "prix", "prezzo", "rrp", "pph", "ppttc", "pricing", "pub price", "rrp value"]
EAN_KW       = ["ean", "isbn", "isbn-13", "isbn13", "ean 13"]
TITLE_KW     = ["title", "titre", "titolo", "libellé article", "libelle article"]
AUTHOR_KW    = ["author", "auteur", "autore", "nom de l'auteur"]
PUBLISHER_KW = ["publisher", "imprint", "editeur", "editore", "marque"]


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def normalize_ean(val) -> str | None:
    """Convert any EAN representation to a 13-digit string, or None."""
    if val is None or val == "":
        return None
    try:
        s = str(int(float(str(val).replace(" ", "").replace("-", ""))))
        return s if len(s) == 13 else None
    except (ValueError, TypeError):
        return None


def parse_filename(filename: str) -> tuple[str | None, float | None]:
    """Extract (country, discount) from a publisher filename."""
    discount_match = re.search(r"(\d+(?:[,\.]\d+)?)\s*%", filename)
    discount = float(discount_match.group(1).replace(",", ".")) / 100 if discount_match else None
    country = None
    if re.search(r"\bGB\b", filename):   country = "GB"
    elif re.search(r"\bFR\b", filename): country = "FR"
    elif re.search(r"\bUS\b", filename): country = "US"
    elif re.search(r"\bDE\b", filename): country = "DE"
    return country, discount


def select_prezzi_key(filename: str) -> str | None:
    """Map a publisher filename to the corresponding Prezzi.xlsx sheet (stripped name)."""
    f = filename.upper()
    if "HACHETTE" in f:   return "HACHETTE 42% FR"
    if "INTERFORUM" in f: return "INTERFORUM 42% FR"
    if "HCUS" in f:       return "HCUS 60% US"
    if "MPS" in f:        return "MPS 58% US"
    if "RHDE" in f:       return "RHDE 51,43% DE"
    if re.search(r"\bDE\b", filename) and not re.search(r"\b(GB|FR|US)\b", filename):
        return "RHDE 51,43% DE"
    if re.search(r"\bGB\b", filename):
        if "60" in f: return "60% GB"
        if "55" in f: return "55% GB"
        if "53" in f: return "53% GB"
        if "45" in f: return "45% GB"
    return None


# ---------------------------------------------------------------------------
# Prezzi.xlsx loader
# ---------------------------------------------------------------------------

def load_prezzi(path) -> dict[str, pd.DataFrame]:
    """
    Load Prezzi.xlsx with data_only=True.
    Accepts a file path (str) or a BytesIO / file-like object.
    Returns dict keyed by stripped sheet name.
    Returns empty dict if file not found (cloud mode).
    """
    import os
    if isinstance(path, str) and not os.path.exists(path):
        return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return {}
    result: dict[str, pd.DataFrame] = {}

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue

        # Row index 2 = headers
        headers = [str(h).strip().lower() if h is not None else "" for h in rows[2]]

        # Lookup price column: prefer "ppttc" (INTERFORUM uses Prix PPTTC), else col 0
        price_col = 0
        for i, h in enumerate(headers):
            if "ppttc" in h:
                price_col = i
                break

        # Find Prezzo IVA column
        prezzo_iva_col = None
        for i, h in enumerate(headers):
            if "prezzo al pubblico" in h:
                prezzo_iva_col = i
                break

        # Find Costo Amazon column
        costo_amazon_col = None
        for i, h in enumerate(headers):
            if "costo netto" in h:
                costo_amazon_col = i
                break

        if prezzo_iva_col is None or costo_amazon_col is None:
            continue

        data = []
        for row in rows[3:]:
            if not row or len(row) <= max(price_col, prezzo_iva_col, costo_amazon_col):
                continue
            if row[price_col] is None:
                continue
            try:
                price = float(row[price_col])
                iva   = float(row[prezzo_iva_col])   if row[prezzo_iva_col]   is not None else None
                amz   = float(row[costo_amazon_col]) if row[costo_amazon_col] is not None else None
                if price > 0:
                    data.append({"price": price, "prezzo_iva": iva, "costo_amazon": amz})
            except (TypeError, ValueError):
                continue

        key = sname.strip()
        if data:
            result[key] = pd.DataFrame(data)

    return result


def lookup_price(
    price,
    prezzi_df: pd.DataFrame | None,
    tolerance: float = 0.20,
) -> tuple[float | None, float | None]:
    """
    Find the row in prezzi_df whose 'price' is closest to the given price.
    Returns (prezzo_iva, costo_amazon).
    If no match within tolerance (20% of price), returns (None, None).
    """
    if prezzi_df is None or price is None:
        return None, None
    try:
        p = float(price)
        if p <= 0:
            return None, None
    except (TypeError, ValueError):
        return None, None

    diffs = (prezzi_df["price"] - p).abs()
    min_diff = diffs.min()

    if pd.isna(min_diff) or min_diff > p * tolerance:
        return None, None

    row = prezzi_df.iloc[diffs.idxmin()]
    return row["prezzo_iva"], row["costo_amazon"]


# ---------------------------------------------------------------------------
# Cache helpers + Web enrichment
# ---------------------------------------------------------------------------

def load_cache() -> dict:
    return st.session_state.get("enrich_cache", {})


def save_cache(cache: dict) -> None:
    st.session_state["enrich_cache"] = cache


def enrich_ean(ean: str, row_data: dict, cache: dict) -> dict:
    """
    Arricchisce i dati per un EAN.
    row_data = dati già estratti dal file editore (pagine, dims, ecc.)
    cache    = dict in session_state per non ripetere chiamate web.

    Strategia: usa prima row_data (file editore), poi web per i vuoti.
    """
    # Partenza dai dati già nel file editore
    result = {
        "description": None,
        "pages":       row_data.get("pagine"),
        "binding":     row_data.get("rilegatura"),
        "language":    row_data.get("lingua"),
        "pub_date":    row_data.get("pub_date"),
        "bisac_code":  row_data.get("bisac"),
        "dim_length":  row_data.get("altezza_mm"),
        "dim_width":   row_data.get("larghezza_mm"),
        "dim_height":  row_data.get("profondita_mm"),
        "weight_kg":   row_data.get("peso_kg"),
    }

    # Campi ancora vuoti che vale la pena cercare online
    need_web = (
        result["description"] is None
        or result["pages"] is None
        or result["language"] is None
        or result["bisac_code"] is None
    )

    if not need_web:
        cache[ean] = result
        return result

    # Check cache web
    web_key = f"web_{ean}"
    if web_key in cache:
        web = cache[web_key]
    else:
        web = _fetch_web(ean)
        cache[web_key] = web

    # Riempi solo i campi ancora vuoti con dati web
    for field in ("description", "pages", "binding", "language",
                  "pub_date", "bisac_code",
                  "dim_length", "dim_width", "dim_height", "weight_kg"):
        if result[field] is None and web.get(field) is not None:
            result[field] = web[field]

    cache[ean] = result
    return result


def _fetch_web(ean: str) -> dict:
    """Chiamate HTTP a Open Library + Google Books. Restituisce dict raw."""
    res = {k: None for k in ("description", "pages", "binding", "language",
                              "pub_date", "bisac_code", "dim_length",
                              "dim_width", "dim_height", "weight_kg")}
    LANG_OL = {
        "/languages/eng": "Inglese", "/languages/fre": "Francese",
        "/languages/ger": "Tedesco", "/languages/ita": "Italiano",
        "/languages/spa": "Spagnolo",
    }
    LANG_GB = {
        "en": "Inglese", "fr": "Francese", "de": "Tedesco",
        "it": "Italiano", "es": "Spagnolo", "pt": "Portogallese",
    }

    # --- Open Library Search (migliore per pagine) ---
    try:
        url = (f"https://openlibrary.org/search.json?isbn={ean}"
               f"&fields=title,number_of_pages_median,physical_format,"
               f"language,first_publish_year,subject")
        req = urllib.request.urlopen(url, timeout=5)
        data = json.loads(req.read().decode())
        docs = data.get("docs", [])
        if docs:
            d = docs[0]
            res["pages"] = d.get("number_of_pages_median")
            pf = d.get("physical_format", "")
            if pf:
                res["binding"] = pf.capitalize()
            langs = d.get("language", [])
            if langs:
                res["language"] = LANG_OL.get(
                    f"/languages/{langs[0]}", langs[0].capitalize()
                )
            subs = d.get("subject", [])
            for s in subs:
                if "/" in s and s[0].isupper():
                    res["bisac_code"] = s
                    break
    except Exception:
        pass
    time.sleep(0.2)

    # --- Open Library /api/books (per descrizione) ---
    try:
        url = (f"https://openlibrary.org/api/books"
               f"?bibkeys=ISBN:{ean}&format=json&jscmd=data")
        req = urllib.request.urlopen(url, timeout=5)
        data = json.loads(req.read().decode())
        book = data.get(f"ISBN:{ean}", {})
        if book:
            if res["pages"] is None:
                pag = book.get("pagination")
                if pag:
                    try:
                        res["pages"] = int(str(pag).replace("p", "").strip())
                    except Exception:
                        pass
            desc = book.get("description")
            if isinstance(desc, dict):
                res["description"] = desc.get("value", "")[:2000]
            elif isinstance(desc, str):
                res["description"] = desc[:2000]
            if res["language"] is None:
                langs = book.get("languages", [])
                if langs:
                    k = langs[0].get("key", "")
                    res["language"] = LANG_OL.get(k, k.split("/")[-1])
            if res["binding"] is None:
                res["binding"] = book.get("physical_format")
    except Exception:
        pass
    time.sleep(0.2)

    # --- Google Books (per descrizione + categoria) ---
    try:
        url = (f"https://www.googleapis.com/books/v1/volumes"
               f"?q=isbn:{ean}&maxResults=1")
        req = urllib.request.urlopen(url, timeout=5)
        data = json.loads(req.read().decode())
        items = data.get("items", [])
        if items:
            info = items[0].get("volumeInfo", {})
            if res["description"] is None:
                d = info.get("description", "")
                res["description"] = d[:2000] if d else None
            if res["pages"] is None:
                pc = info.get("pageCount", 0)
                res["pages"] = pc if pc and pc > 0 else None
            if res["language"] is None:
                res["language"] = LANG_GB.get(
                    info.get("language", ""), info.get("language")
                )
            if res["pub_date"] is None:
                res["pub_date"] = info.get("publishedDate")
            if res["bisac_code"] is None:
                cats = info.get("categories", [])
                if cats:
                    res["bisac_code"] = cats[0]
            dims = info.get("dimensions", {})
            if dims:
                def to_mm(s):
                    try:
                        return round(float(str(s).replace(" inches", "").strip()) * 25.4)
                    except Exception:
                        return None
                if res["dim_length"] is None:
                    res["dim_length"] = to_mm(dims.get("height"))
                if res["dim_width"] is None:
                    res["dim_width"] = to_mm(dims.get("width"))
                if res["dim_height"] is None:
                    res["dim_height"] = to_mm(dims.get("thickness"))
    except Exception:
        pass
    time.sleep(0.2)

    return res


# ---------------------------------------------------------------------------
# Shared row-extraction helper
# ---------------------------------------------------------------------------

def _rows_from_df(
    df: pd.DataFrame,
    ean_col: str,
    title_col: str | None,
    author_col: str | None,
    publisher_col: str | None,
    price_col: str | None,
    default_publisher: str = "",
) -> list[dict]:
    """Extract normalised rows from a DataFrame using named columns."""
    rows = []
    for _, row in df.iterrows():
        ean = normalize_ean(row.get(ean_col)) if ean_col in df.columns else None
        if not ean:
            continue

        def _str(col):
            if col and col in df.columns:
                return str(row.get(col) or "").strip()
            return ""

        publisher = _str(publisher_col) or default_publisher
        price = None
        if price_col and price_col in df.columns:
            raw = row.get(price_col)
            if raw is not None:
                try:
                    price = float(raw)
                except (TypeError, ValueError):
                    pass

        rows.append({
            "ean":           ean,
            "titolo":        _str(title_col),
            "editore":       publisher,
            "autore":        _str(author_col),
            "prezzo":        price,
            "pagine":        None,
            "altezza_mm":    None,
            "larghezza_mm":  None,
            "profondita_mm": None,
            "peso_kg":       None,
            "rilegatura":    None,
            "bisac":         None,
            "pub_date":      None,
            "lingua":        None,
        })
    return rows


# ---------------------------------------------------------------------------
# Publisher-specific extractors
# ---------------------------------------------------------------------------

def extract_gb60_canongate(data: bytes) -> tuple[list[dict], list[str]]:
    """GB 60% Canongate — sheet 'Pub schedule 26 ', header row 0."""
    rows, warns = [], []
    BIND_MAP = {"TPB": "Trade Paperback", "HB": "Hardcover",
                "MM": "Mass Market Paperback", "PB": "Paperback",
                "BB": "Board Book", "CD": "Hardcover"}

    def parse_mm(v):
        if v is None:
            return None
        try:
            return int(str(v).replace("mm", "").strip())
        except Exception:
            return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data))
        sheet_name = next(
            (s for s in wb.sheetnames if "pub schedule" in s.lower()), None
        )
        if not sheet_name:
            return [], ["Sheet 'Pub schedule' not found in Canongate file"]

        df = pd.read_excel(io.BytesIO(data), sheet_name=sheet_name, header=0)
        df.columns = [str(c).strip() for c in df.columns]
        rows = _rows_from_df(df, "ISBN-13", "Title", "Author", "Imprint", "Pub Price")

        # Build EAN -> df row lookup for extra fields
        ean_to_dfrow = {}
        for _, dfrow in df.iterrows():
            e = normalize_ean(dfrow.get("ISBN-13"))
            if e:
                ean_to_dfrow[e] = dfrow

        for r in rows:
            dfrow = ean_to_dfrow.get(r["ean"], {})
            r["lingua"] = "Inglese"
            if "Edition Extent" in df.columns:
                try:
                    r["pagine"] = int(dfrow.get("Edition Extent"))
                except Exception:
                    pass
            r["altezza_mm"]    = parse_mm(dfrow.get("Page Height"))    if "Page Height" in df.columns else None
            r["larghezza_mm"]  = parse_mm(dfrow.get("Page Width"))     if "Page Width" in df.columns else None
            r["profondita_mm"] = parse_mm(dfrow.get("Depth"))          if "Depth" in df.columns else None
            if "Edition Weight (Grams)" in df.columns:
                w = dfrow.get("Edition Weight (Grams)")
                try:
                    r["peso_kg"] = round(float(w) / 1000, 3)
                except Exception:
                    pass
            if "Binding Initials" in df.columns:
                code = dfrow.get("Binding Initials")
                r["rilegatura"] = BIND_MAP.get(str(code).strip().upper(), str(code).strip()) if code else None
            if "Edition Group Codes: BISAC" in df.columns:
                bisac_raw = dfrow.get("Edition Group Codes: BISAC")
                r["bisac"] = str(bisac_raw).split(",")[0].strip() if bisac_raw else None
            if "Pub Date" in df.columns:
                r["pub_date"] = dfrow.get("Pub Date")
    except Exception as exc:
        warns.append(f"Canongate extract error: {exc}")
    return rows, warns


def extract_gb45_bloomsbury_academic(data: bytes) -> tuple[list[dict], list[str]]:
    """GB 45% Bloomsbury Academic — sheet 'Order Form', header row 0."""
    rows, warns = [], []
    try:
        df = pd.read_excel(io.BytesIO(data), sheet_name="Order Form", header=0)
        df.columns = [str(c).strip() for c in df.columns]
        rows = _rows_from_df(
            df, "ISBN", "Title", None, None, "Price",
            default_publisher="Bloomsbury",
        )
        ean_to_dfrow = {}
        for _, dfrow in df.iterrows():
            e = normalize_ean(dfrow.get("ISBN"))
            if e:
                ean_to_dfrow[e] = dfrow
        for r in rows:
            dfrow = ean_to_dfrow.get(r["ean"], {})
            r["lingua"] = "Inglese"
            if "Publication Date" in df.columns:
                r["pub_date"] = dfrow.get("Publication Date")
    except Exception as exc:
        warns.append(f"Bloomsbury Academic extract error: {exc}")
    return rows, warns


def extract_gb53_lonely_planet(data: bytes) -> tuple[list[dict], list[str]]:
    """
    GB 53% Lonely Planet — sheet 'CDPubListGBPResults'
    No header row: col 0 = title, col 3 = ISBN, col 4 = RRP Value
    """
    rows, warns = [], []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        sheet_name = next(
            (s for s in wb.sheetnames if "cdpub" in s.lower()), wb.sheetnames[0]
        )
        ws = wb[sheet_name]

        for row in ws.iter_rows(values_only=True):
            if len(row) < 5:
                continue
            ean = normalize_ean(row[3])
            if not ean:
                continue
            price = None
            if row[4] is not None:
                try:
                    price = float(row[4])
                except (TypeError, ValueError):
                    pass
            pagine = None
            if len(row) > 7 and row[7] is not None:
                try:
                    pagine = int(row[7])
                except Exception:
                    pass
            peso = None
            if len(row) > 10 and row[10] is not None:
                try:
                    peso = float(str(row[10]).strip())
                except Exception:
                    pass
            rileg = None
            if len(row) > 11 and row[11]:
                pf = str(row[11]).strip()
                rileg = "Paperback" if "paperback" in pf.lower() or "softback" in pf.lower() else pf
            rows.append({
                "ean":           ean,
                "titolo":        str(row[0] or "").strip(),
                "editore":       "Lonely Planet",
                "autore":        "",
                "prezzo":        price,
                "pagine":        pagine,
                "altezza_mm":    None,
                "larghezza_mm":  None,
                "profondita_mm": None,
                "peso_kg":       peso,
                "rilegatura":    rileg,
                "bisac":         None,
                "pub_date":      None,
                "lingua":        "Inglese",
            })
    except Exception as exc:
        warns.append(f"Lonely Planet extract error: {exc}")
    return rows, warns


def extract_gb55_bloomsbury_italian(data: bytes) -> tuple[list[dict], list[str]]:
    """GB 55% Bloomsbury Italian Interest — sheet 'Italian Interest', header row 3."""
    rows, warns = [], []
    try:
        df = pd.read_excel(io.BytesIO(data), sheet_name="Italian Interest", header=3)
        df.columns = [str(c).strip() for c in df.columns]
        rows = _rows_from_df(df, "ISBN", "Title", "Author", "Publisher", "Price")
        ean_to_dfrow = {}
        for _, dfrow in df.iterrows():
            e = normalize_ean(dfrow.get("ISBN"))
            if e:
                ean_to_dfrow[e] = dfrow
        for r in rows:
            dfrow = ean_to_dfrow.get(r["ean"], {})
            r["lingua"] = "Inglese"
            if "Format" in df.columns:
                r["rilegatura"] = dfrow.get("Format")
            if "Category" in df.columns:
                r["bisac"] = dfrow.get("Category")
    except Exception as exc:
        warns.append(f"Bloomsbury Italian extract error: {exc}")
    return rows, warns


def extract_hachette(data: bytes) -> tuple[list[dict], list[str]]:
    """HACHETTE 42% FR — sheet 'Bon de commande', header row 0."""
    rows, warns = [], []
    try:
        df = pd.read_excel(io.BytesIO(data), sheet_name="Bon de commande", header=0)
        df.columns = [str(c).strip() for c in df.columns]
        # 'Titre' is the real title (5th col); 'Titre intercalaire' is the section label
        rows = _rows_from_df(
            df, "EAN 13", "Titre", "Auteur", "Marque \u00e9ditoriale", "Prix"
        )
        ean_to_dfrow = {}
        for _, dfrow in df.iterrows():
            e = normalize_ean(dfrow.get("EAN 13"))
            if e:
                ean_to_dfrow[e] = dfrow
        for r in rows:
            dfrow = ean_to_dfrow.get(r["ean"], {})
            r["lingua"] = "Francese"
            if "Date de MEV" in df.columns:
                r["pub_date"] = dfrow.get("Date de MEV")
    except Exception as exc:
        warns.append(f"Hachette extract error: {exc}")
    return rows, warns


def extract_interforum(data: bytes) -> tuple[list[dict], list[str]]:
    """INTERFORUM 42% FR — sheet 'Foglio1', header row 0."""
    rows, warns = [], []
    try:
        df = pd.read_excel(io.BytesIO(data), sheet_name="Foglio1", header=0)
        df.columns = [str(c).strip() for c in df.columns]
        rows = _rows_from_df(
            df, "EAN", "Libell\u00e9 Article", "Nom de l'Auteur", "Editeur", "Prix PPTTC"
        )
        for r in rows:
            r["lingua"] = "Francese"
    except Exception as exc:
        warns.append(f"Interforum extract error: {exc}")
    return rows, warns


def extract_hcus(data: bytes) -> tuple[list[dict], list[str]]:
    """
    HCUS 60% US (HarperCollins) — 3 sheets: May/April/March 2026
    Header at row 14 (0-indexed); price is at column index 7.
    """
    rows, warns = [], []
    target_sheets = ["May 2026", "April 2026", "March 2026"]
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        available = [s for s in target_sheets if s in wb.sheetnames]

        for sname in available:
            ws = wb[sname]
            all_rows = list(ws.iter_rows(values_only=True))

            if len(all_rows) <= 15:
                continue

            # Data starts at row index 15 (header is row 14)
            for row in all_rows[15:]:
                if len(row) < 8:
                    continue
                ean = normalize_ean(row[2])   # ISBN-13 at col index 2
                if not ean:
                    continue

                title  = str(row[3] or "").strip()
                author = str(row[4] or "").strip()

                price_val = row[7]            # Price at col index 7
                if price_val is None:
                    continue
                try:
                    price = float(price_val)
                except (TypeError, ValueError):
                    continue

                rileg = str(row[5] or "").strip() if len(row) > 5 else None
                pub_dt = row[6] if len(row) > 6 else None
                bisac_val = str(row[9] or "").strip() if len(row) > 9 else None
                rows.append({
                    "ean":           ean,
                    "titolo":        title,
                    "editore":       "HarperCollins",
                    "autore":        author,
                    "prezzo":        price,
                    "pagine":        None,
                    "altezza_mm":    None,
                    "larghezza_mm":  None,
                    "profondita_mm": None,
                    "peso_kg":       None,
                    "rilegatura":    rileg or None,
                    "bisac":         bisac_val or None,
                    "pub_date":      pub_dt,
                    "lingua":        "Inglese",
                })
    except Exception as exc:
        warns.append(f"HCUS extract error: {exc}")
    return rows, warns


def extract_mps(data: bytes) -> tuple[list[dict], list[str]]:
    """MPS 58% US (Macmillan) — sheet 'ADULT', header row 2."""
    rows, warns = [], []
    try:
        df = pd.read_excel(io.BytesIO(data), sheet_name="ADULT", header=2)
        df.columns = [str(c).strip() for c in df.columns]
        rows = _rows_from_df(df, "ISBN", "Title", "Author", "Imprint", "Price")
        ean_to_dfrow = {}
        for _, dfrow in df.iterrows():
            e = normalize_ean(dfrow.get("ISBN"))
            if e:
                ean_to_dfrow[e] = dfrow
        for r in rows:
            dfrow = ean_to_dfrow.get(r["ean"], {})
            r["lingua"] = "Inglese"
            if "Format" in df.columns:
                r["rilegatura"] = dfrow.get("Format")
            if "Pub Date" in df.columns:
                r["pub_date"] = dfrow.get("Pub Date")
    except Exception as exc:
        warns.append(f"MPS extract error: {exc}")
    return rows, warns


# ---------------------------------------------------------------------------
# Auto-detection fallback
# ---------------------------------------------------------------------------

def extract_auto(data: bytes, filename: str) -> tuple[list[dict], list[str]]:
    """
    Generic fallback parser: scans first 20 rows to find the header row
    using keyword matching, then extracts EAN/title/author/publisher/price.
    """
    rows, warns = [], []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        all_kw = PRICE_KW + EAN_KW + TITLE_KW + AUTHOR_KW + PUBLISHER_KW

        for sname in wb.sheetnames:
            ws = wb[sname]
            sheet_rows = list(ws.iter_rows(values_only=True))
            if not sheet_rows:
                continue

            def score(r):
                text = " ".join(str(c).lower().strip() for c in r if c)
                return sum(1 for kw in all_kw if kw in text)

            scan = min(20, len(sheet_rows))
            best = max(range(scan), key=lambda i: score(sheet_rows[i]))
            if score(sheet_rows[best]) < 2:
                continue

            headers = [
                str(c).strip().lower() if c is not None else ""
                for c in sheet_rows[best]
            ]

            def find(kws):
                for i, h in enumerate(headers):
                    if any(kw in h for kw in kws):
                        return i
                return None

            ean_idx   = find(EAN_KW)
            title_idx = find(TITLE_KW)
            price_idx = find(PRICE_KW)
            auth_idx  = find(AUTHOR_KW)
            pub_idx   = find(PUBLISHER_KW)

            if ean_idx is None:
                warns.append(
                    f"EAN column not detected in sheet '{sname}' of {filename} — skipped"
                )
                continue

            default_pub = filename.rsplit(".", 1)[0]

            for row in sheet_rows[best + 1:]:
                if len(row) <= ean_idx:
                    continue
                ean = normalize_ean(row[ean_idx])
                if not ean:
                    continue

                def _get(idx):
                    if idx is not None and len(row) > idx:
                        return str(row[idx] or "").strip()
                    return ""

                price = None
                if price_idx is not None and len(row) > price_idx and row[price_idx] is not None:
                    try:
                        price = float(row[price_idx])
                    except (TypeError, ValueError):
                        pass

                publisher = _get(pub_idx) or default_pub

                rows.append({
                    "ean":     ean,
                    "titolo":  _get(title_idx),
                    "editore": publisher,
                    "autore":  _get(auth_idx),
                    "prezzo":  price,
                })
    except Exception as exc:
        warns.append(f"Auto-detect error in {filename}: {exc}")
    return rows, warns


# ---------------------------------------------------------------------------
# Extractor dispatcher
# ---------------------------------------------------------------------------

def get_extractor(filename: str):
    """Return the hardcoded extractor for a known publisher file, or None for auto-detect."""
    f = filename.upper()
    if "CANONGATE" in f:                                return extract_gb60_canongate
    if "BLOOMSBURY ACADEMIC" in f:                      return extract_gb45_bloomsbury_academic
    if "LONELY PLANET" in f:                            return extract_gb53_lonely_planet
    if "BLOOMSBURY ITALIAN" in f or "ITALIAN INTEREST" in f: return extract_gb55_bloomsbury_italian
    if "HACHETTE" in f:                                 return extract_hachette
    if "INTERFORUM" in f:                               return extract_interforum
    if "HCUS" in f:                                     return extract_hcus
    if "MPS" in f:                                      return extract_mps
    # Fallback by country+discount pattern
    if re.search(r"\bGB\b", filename):
        if "60" in f: return extract_gb60_canongate
        if "55" in f: return extract_gb55_bloomsbury_italian
        if "53" in f: return extract_gb53_lonely_planet
        if "45" in f: return extract_gb45_bloomsbury_academic
    return None


# ---------------------------------------------------------------------------
# Extraction (step 1) — no web calls, returns base DataFrame
# ---------------------------------------------------------------------------

def process_files(
    uploaded_files, prezzi_db: dict
) -> tuple[pd.DataFrame, list[str], list[str]]:
    """
    Parse all uploaded files and do Prezzi price lookups.
    Accetta sia lista di UploadedFile (locale) sia dict {fname: bytes} (cloud).
    Saves rows_raw in st.session_state["rows_raw"].
    Returns (df_base, warnings, warning_eans_list).
    """
    all_base:    list[dict] = []
    all_raw:     list[dict] = []
    all_warnings: list[str] = []
    missing_price_eans: set[str] = set()

    # Normalizza input: accetta sia dict {nome: bytes} sia lista di UploadedFile
    if isinstance(uploaded_files, dict):
        files_iter = [(fname, data) for fname, data in uploaded_files.items()]
    else:
        files_iter = [(f.name, f.read()) for f in uploaded_files]

    for fname, data in files_iter:

        paese, sconto = parse_filename(fname)
        prezzi_key    = select_prezzi_key(fname)
        prezzi_df     = prezzi_db.get(prezzi_key) if prezzi_key else None

        if prezzi_key and prezzi_df is None:
            all_warnings.append(
                f"⚠️ Foglio Prezzi '{prezzi_key}' non trovato per {fname} — colonne G-H vuote"
            )

        extractor = get_extractor(fname)
        if extractor:
            extracted, warns = extractor(data)
        else:
            extracted, warns = extract_auto(data, fname)
            if not extracted:
                all_warnings.append(f"⚠️ Struttura non riconosciuta per {fname} — saltato")

        all_warnings.extend(warns)

        for r in extracted:
            all_raw.append(r)  # salva dati grezzi con extra fields per enrichment
            prezzo_iva, costo_amazon = lookup_price(r.get("prezzo"), prezzi_df)
            ean = r["ean"]
            if prezzo_iva is None and r.get("prezzo") is not None:
                missing_price_eans.add(ean)
            all_base.append({
                "EAN":     ean,
                "TITOLO":  r.get("titolo", ""),
                "EDITORE": r.get("editore", ""),
                "PAESE":   paese or "",
                "SCONTO":  sconto,
                "Prezzo Editore": r.get("prezzo"),
                "Prezzo al pubblico consigliato (IVA inclusa)": prezzo_iva,
                "Costo netto di cessione ad Amazon":            costo_amazon,
            })

    st.session_state["rows_raw"] = all_raw

    df_base = pd.DataFrame(all_base, columns=HEADERS_BASE)
    df_base.drop_duplicates(subset="EAN", keep="first", inplace=True)
    df_base.reset_index(drop=True, inplace=True)

    warning_eans_list = list(missing_price_eans & set(df_base["EAN"].tolist()))
    return df_base, all_warnings, warning_eans_list


def build_enriched_df(df_base: pd.DataFrame, enriched_data: dict) -> pd.DataFrame:
    """
    Build the full 27-column DataFrame from df_base + enriched_data.
    enriched_data = {ean: {description, pages, binding, ...}}
    Fixed Amazon fields are filled from FIXED_VALUES.
    """
    rows = []
    for _, row in df_base.iterrows():
        ean = str(row["EAN"])
        enr = enriched_data.get(ean, {})
        new_row = row.to_dict()
        new_row["Descrizione prodotto"]    = enr.get("description")
        new_row["Numero di pagine"]        = enr.get("pages")
        new_row["Rilegatura"]              = enr.get("binding")
        new_row["Lingua"]                  = enr.get("language")
        new_row["Data di pubblicazione"]   = enr.get("pub_date")
        new_row["Codice soggetto (BISAC)"] = enr.get("bisac_code")
        new_row["Dimensione L (mm)"]       = enr.get("dim_length")
        new_row["Dimensione W (mm)"]       = enr.get("dim_width")
        new_row["Dimensione H (mm)"]       = enr.get("dim_height")
        new_row["Peso (kg)"]               = enr.get("weight_kg")
        for col, val in FIXED_VALUES.items():
            new_row[col] = val
        rows.append(new_row)
    all_cols = HEADERS_BASE + HEADERS_WEB + HEADERS_FIXED
    return pd.DataFrame(rows, columns=all_cols)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_xlsx(df: pd.DataFrame,
                warning_eans: list[str] | None = None,
                enriched: bool = False) -> bytes:
    """
    Write df to a formatted .xlsx file and return the bytes.
    Header colours by column group (based on HEADERS_BASE/WEB/FIXED membership):
      - HEADERS_BASE  : yellow  #FFD700
      - HEADERS_WEB   : light blue #DDEEFF
      - HEADERS_FIXED : light green #DDFFD8
    Rows with no price match: light red fill on HEADERS_BASE cols only.
    """
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foglio1"

    YELLOW = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    BLUE   = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    GREEN  = PatternFill(start_color="DDFFD8", end_color="DDFFD8", fill_type="solid")
    WARN   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    bold   = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    warn_set = set(warning_eans or [])
    cols = list(df.columns)

    base_set  = set(HEADERS_BASE)
    web_set   = set(HEADERS_WEB)
    fixed_set = set(HEADERS_FIXED)

    # --- Header row ---
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = bold
        cell.alignment = center
        if col in base_set:   cell.fill = YELLOW
        elif col in web_set:  cell.fill = BLUE
        elif col in fixed_set: cell.fill = GREEN

    # --- Data rows ---
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        is_warn = str(row.get("EAN", "")) in warn_set
        for ci, col in enumerate(cols, 1):
            val  = row.get(col)
            cell = ws.cell(row=ri, column=ci)

            if col == "EAN":
                cell.value = str(val) if val else ""
                cell.number_format = "@"
            elif col == "SCONTO":
                if val is not None:
                    try:
                        cell.value = float(val)
                        cell.number_format = "0%"
                    except (TypeError, ValueError):
                        cell.value = val
                else:
                    cell.value = val
            elif col in ("Prezzo Editore",
                         "Prezzo al pubblico consigliato (IVA inclusa)",
                         "Costo netto di cessione ad Amazon"):
                if val is not None:
                    try:
                        cell.value = float(val)
                        cell.number_format = "#,##0.00"
                    except (TypeError, ValueError):
                        cell.value = val
                else:
                    cell.value = val
            elif col in ("Numero di pagine", "Dimensione L (mm)",
                         "Dimensione W (mm)", "Dimensione H (mm)"):
                if val is not None:
                    try:
                        cell.value = int(val)
                    except (TypeError, ValueError):
                        cell.value = val
                else:
                    cell.value = val
            elif col == "Peso (kg)":
                if val is not None:
                    try:
                        cell.value = float(val)
                        cell.number_format = "#,##0.000"
                    except (TypeError, ValueError):
                        cell.value = val
                else:
                    cell.value = val
            else:
                cell.value = val

            if is_warn and col in base_set:
                cell.fill = WARN

    # --- Auto column width ---
    for ci, col in enumerate(cols, 1):
        col_letter = get_column_letter(ci)
        max_len = len(col)
        limit = 60 if col == "Descrizione prodotto" else 35
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=ci).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, limit)

    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------


def main():
    st.set_page_config(
        page_title="Johnsons Books",
        page_icon="📚",
        layout="wide",
    )

    st.title("📚 Johnsons Books — Publisher File Extractor")
    st.caption("Carica uno o più file editori per generare il file di output unificato")

    # ----------------------------------------------------------------
    # SIDEBAR — Tabella Prezzi (uploader separato, non interferisce)
    # ----------------------------------------------------------------
    with st.sidebar:
        st.header("⚙️ Tabella Prezzi")

        # Inizializza prezzi da file locale se disponibile
        if "prezzi_db" not in st.session_state:
            local = load_prezzi(PREZZI_PATH)
            st.session_state["prezzi_db"] = local if local else {}

        prezzi_file = st.file_uploader(
            "Sostituisci Prezzi.xlsx",
            type=["xlsx"],
            key="prezzi_uploader",
        )
        if prezzi_file is not None:
            b = prezzi_file.getvalue()
            if b:
                parsed = load_prezzi(io.BytesIO(b))
                if parsed:
                    st.session_state["prezzi_db"] = parsed
                    st.success(f"✅ Caricato ({len(parsed)} fogli)")
                else:
                    st.error("File non valido")

        prezzi_db = st.session_state["prezzi_db"]
        if prezzi_db:
            st.caption(f"Attivo: {len(prezzi_db)} fogli")
        else:
            st.warning("Carica Prezzi.xlsx per abilitare il calcolo prezzi.")

    prezzi_db = st.session_state.get("prezzi_db", {})

    # ----------------------------------------------------------------
    # SESSION STATE
    # ----------------------------------------------------------------
    for k, v in [
        ("df_base", None),
        ("df_enriched", None),
        ("enrich_cache", {}),
        ("warnings_base", []),
        ("warning_eans", []),
        ("rows_raw", []),
        ("uploaded_bytes", {}),
    ]:
        if k not in st.session_state:
            st.session_state[k] = v

    # ----------------------------------------------------------------
    # BODY — File uploader editori (unico, senza expander)
    # ----------------------------------------------------------------
    uploaded_files = st.file_uploader(
        "Carica file editori (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Puoi trascinare più file contemporaneamente",
    )

    if not uploaded_files:
        st.info("👆 Carica uno o più file editori per iniziare")
        return

    # Leggi i bytes UNA SOLA VOLTA e salvali in session_state
    # Confronta per nomi file — se cambiano, rileggi
    current_names = tuple(sorted(f.name for f in uploaded_files))
    saved_names   = tuple(sorted(st.session_state["uploaded_bytes"].keys()))

    if current_names != saved_names:
        new_bytes = {f.name: f.getvalue() for f in uploaded_files if f.getvalue()}
        if new_bytes:
            st.session_state["uploaded_bytes"] = new_bytes
            st.session_state["df_base"]        = None
            st.session_state["df_enriched"]    = None

    uploaded_bytes = st.session_state["uploaded_bytes"]

    if not uploaded_bytes:
        st.warning("⚠️ Nessun file leggibile. Ricarica i file editori.")
        return

    # ----------------------------------------------------------------
    # RIEPILOGO FILE
    # ----------------------------------------------------------------
    st.subheader(f"{len(uploaded_bytes)} file selezionati")
    summary_rows = []
    for fname in uploaded_bytes:
        paese, sconto = parse_filename(fname)
        key           = select_prezzi_key(fname)
        ext           = get_extractor(fname)
        summary_rows.append({
            "File":          fname,
            "Paese":         paese or "?",
            "Sconto":        f"{sconto:.0%}" if sconto else "?",
            "Foglio Prezzi": key or "—",
            "Parser":        ext.__name__ if ext else "auto-detect",
        })
    st.dataframe(
        pd.DataFrame(summary_rows),
        width="stretch",
        hide_index=True,
    )

    # ----------------------------------------------------------------
    # BOTTONE 1 — Output Base
    # ----------------------------------------------------------------
    if st.button("🔄 Genera Output Base", type="primary", width="stretch"):
        with st.spinner("Estrazione dati dai file editori…"):
            df_base, warnings, warning_eans = process_files(uploaded_bytes, prezzi_db)

        if df_base.empty:
            st.error("Nessuna riga estratta. Verifica i file caricati.")
            for w in warnings:
                st.warning(w)
        else:
            st.session_state["df_base"]       = df_base
            st.session_state["df_enriched"]   = None
            st.session_state["warnings_base"] = warnings
            st.session_state["warning_eans"]  = warning_eans

    # ----------------------------------------------------------------
    # RISULTATO BASE
    # ----------------------------------------------------------------
    if st.session_state["df_base"] is not None:
        df        = st.session_state["df_base"]
        warnings  = st.session_state["warnings_base"]
        warn_eans = st.session_state["warning_eans"]

        col1, col2, col3 = st.columns(3)
        col1.metric("File processati", len(uploaded_bytes))
        col2.metric("EAN totali",      len(df))
        col3.metric("Senza prezzo",    len(warn_eans))

        if warnings:
            with st.expander(f"⚠️ {len(warnings)} avvisi"):
                for w in warnings:
                    st.warning(w)

        st.subheader("Anteprima output base")
        st.dataframe(df.head(50), width="stretch", hide_index=True)

        xlsx_base = export_xlsx(df, warn_eans, enriched=False)
        st.download_button(
            label="📥 Download output_base.xlsx",
            data=xlsx_base,
            file_name="output_base.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

        st.divider()

        # ----------------------------------------------------------------
        # BOTTONE 2 — Arricchimento Web
        # ----------------------------------------------------------------
        n_ean        = len(df)
        cache        = st.session_state["enrich_cache"]
        cached_count = sum(1 for e in df["EAN"].tolist() if f"web_{e}" in cache)
        da_cercare   = n_ean - cached_count
        mins_est     = max(1, round(da_cercare * 0.7 / 60))

        st.subheader("🌐 Arricchimento dati (opzionale)")
        st.caption(
            "Aggiunge descrizione, pagine, rilegatura, lingua, dimensioni, BISAC. "
            "Prima usa i dati dal file editore, poi cerca online i campi mancanti."
        )
        if cached_count > 0:
            st.caption(f"✅ {cached_count} EAN già in cache — {da_cercare} nuove ricerche web")

        if st.button(
            f"🌐 Arricchisci dati (~{mins_est} min per {da_cercare} ricerche web)",
            width="stretch",
        ):
            rows_raw      = st.session_state.get("rows_raw", [])
            eans          = df["EAN"].tolist()
            progress_bar  = st.progress(0, text="Arricchimento in corso…")
            enriched_data = {}

            for i, ean in enumerate(eans):
                row_data           = next((r for r in rows_raw if r.get("ean") == ean), {})
                enriched_data[ean] = enrich_ean(ean, row_data, cache)
                progress_bar.progress(
                    (i + 1) / len(eans),
                    text=f"EAN {i+1}/{len(eans)}…",
                )

            st.session_state["enrich_cache"] = cache
            progress_bar.empty()
            df_enriched = build_enriched_df(df, enriched_data)
            st.session_state["df_enriched"] = df_enriched
            st.success(f"✅ Completato per {len(eans)} EAN")

        if st.session_state["df_enriched"] is not None:
            df_e = st.session_state["df_enriched"]
            st.subheader("Anteprima output completo")
            st.dataframe(df_e.head(50), width="stretch", hide_index=True)

            xlsx_enriched = export_xlsx(df_e, warn_eans, enriched=True)
            st.download_button(
                label="📥 Download output_completo.xlsx",
                data=xlsx_enriched,
                file_name="output_completo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                width="stretch",
            )


if __name__ == "__main__":
    main()
